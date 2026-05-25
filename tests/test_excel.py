"""Tests for src/scanner/report/excel.py.

All assertions are structural (sheet presence, row/column counts, cell values,
null rendering). Styling is deliberately not tested — it's a presentation
concern that changes without breaking the contract.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from scanner.report.excel import build_excel_workbook, write_excel

# ── Fixtures ─────────────────────────────────────────────────────────────────

_REQUIRED_HEADERS = {
    "Ticker", "Exchange", "Name", "Direction",
    "Combo Score", "Rank Score", "Agreement",
    "Signals Firing", "Vol Confirmation", "Volume Confirmation",
    "Days Since Breakout", "Alignment Fraction",
    "Sector", "Region", "Market Cap (USD)",
}


def _sample_dict(n_tickers: int = 3) -> dict:
    """Build a minimal canonical dashboard dict with n tickers."""
    tickers = []
    for i in range(n_tickers):
        rank = 0.90 - i * 0.10
        direction = ["buy", "sell", "neutral"][i % 3]
        tickers.append({
            "ticker": f"TK{i}",
            "exchange": "US",
            "date": "2024-06-15",
            "direction": direction,
            "combo_score": 0.30 + i * 0.05,
            "rank_score": rank,
            "agreement_count": 8 - i,
            "n_trade_indicators": 8,
            "signals_firing": ["rsi", "stochastic"] if direction != "neutral" else [],
            "vol_confirmation": "confirm",
            "volume_confirmation": "confirm" if i < 2 else "reject",
            "days_since_breakout": i if i < 2 else None,
            "meta": {
                "name": f"Company {i}" if i < 2 else None,
                "currency": "USD",
                "market_cap_usd": 1e12 - i * 1e11 if i < 2 else None,
                "sector": "Technology" if i == 0 else None,
                "region": "North America" if i == 0 else None,
            },
            "mtf_alignment": {
                "resolutions_available": 2 if i == 0 else 0,
                "resolutions_aligned": 2 if i == 0 else 0,
                "alignment_fraction": 1.0 if i == 0 else 0.0,
            },
            "indicators": {},
        })
    return {
        "meta": {
            "schema_version": "1.0",
            "run_date": "2024-06-15",
            "generated_at": "2024-06-15T12:00:00+00:00",
            "scope": "sample",
            "combination_name": "default",
            "n_tickers_scored": n_tickers,
            "n_tickers_universe": 100,
            "n_buy": sum(1 for t in tickers if t["direction"] == "buy"),
            "n_sell": sum(1 for t in tickers if t["direction"] == "sell"),
            "n_neutral": sum(1 for t in tickers if t["direction"] == "neutral"),
        },
        "tickers": tickers,
    }


def _empty_dict() -> dict:
    return {
        "meta": {
            "schema_version": "1.0",
            "run_date": "2024-06-15",
            "generated_at": "2024-06-15T12:00:00+00:00",
            "scope": "sample",
            "combination_name": "default",
            "n_tickers_scored": 0,
            "n_tickers_universe": 0,
            "n_buy": 0,
            "n_sell": 0,
            "n_neutral": 0,
        },
        "tickers": [],
    }


# ── Tests ────────────────────────────────────────────────────────────────────


class TestSheetPresent:
    def test_ranked_summary_sheet_exists(self) -> None:
        wb = build_excel_workbook(_sample_dict())
        assert "Ranked Summary" in wb.sheetnames

    def test_single_sheet_only(self) -> None:
        wb = build_excel_workbook(_sample_dict())
        assert len(wb.sheetnames) == 1


class TestRowCount:
    def test_row_count_matches_n_tickers(self) -> None:
        data = _sample_dict(n_tickers=5)
        wb = build_excel_workbook(data)
        ws = wb["Ranked Summary"]
        # 1 header + 5 data rows
        assert ws.max_row == 6

    def test_row_count_three_tickers(self) -> None:
        wb = build_excel_workbook(_sample_dict(3))
        ws = wb["Ranked Summary"]
        assert ws.max_row == 4


class TestRankedOrder:
    def test_rank_score_descending(self) -> None:
        data = _sample_dict(5)
        wb = build_excel_workbook(data)
        ws = wb["Ranked Summary"]
        # Find rank_score column
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        rank_col = headers.index("Rank Score") + 1
        scores = [ws.cell(row=r, column=rank_col).value for r in range(2, ws.max_row + 1)]
        assert scores == sorted(scores, reverse=True)

    def test_first_ticker_has_highest_rank(self) -> None:
        data = _sample_dict(3)
        wb = build_excel_workbook(data)
        ws = wb["Ranked Summary"]
        # First data row should be TK0 (rank 0.90)
        assert ws.cell(row=2, column=1).value == "TK0"


class TestRequiredColumns:
    def test_all_required_headers_present(self) -> None:
        wb = build_excel_workbook(_sample_dict())
        ws = wb["Ranked Summary"]
        headers = {ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)}
        assert _REQUIRED_HEADERS <= headers


class TestNullMeta:
    def test_null_meta_renders_as_blank(self) -> None:
        """Missing meta fields should render as blank (None), not 'None' or 'nan'."""
        data = _sample_dict(3)
        wb = build_excel_workbook(data)
        ws = wb["Ranked Summary"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        name_col = headers.index("Name") + 1
        sector_col = headers.index("Sector") + 1
        region_col = headers.index("Region") + 1
        mcap_col = headers.index("Market Cap (USD)") + 1

        # TK2 (row 4) has all-null meta
        for col in (name_col, sector_col, region_col, mcap_col):
            cell_value = ws.cell(row=4, column=col).value
            assert cell_value is None, (
                f"Column {headers[col-1]} row 4 should be blank, got {cell_value!r}"
            )

    def test_no_none_string_in_cells(self) -> None:
        """No cell should contain the literal strings 'None' or 'nan'."""
        data = _sample_dict(3)
        wb = build_excel_workbook(data)
        ws = wb["Ranked Summary"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            for cell_value in row:
                if isinstance(cell_value, str):
                    assert cell_value.lower() not in ("none", "nan"), (
                        f"Found literal {cell_value!r} in a cell"
                    )


class TestEmptyRun:
    def test_empty_run_valid_workbook(self) -> None:
        wb = build_excel_workbook(_empty_dict())
        ws = wb["Ranked Summary"]
        # Header row only
        assert ws.max_row == 1
        assert ws.cell(row=1, column=1).value == "Ticker"


class TestWriteExcel:
    def test_write_creates_file(self, tmp_path: Path) -> None:
        out = write_excel(_sample_dict(), tmp_path / "out.xlsx")
        assert out.exists()
        assert out.suffix == ".xlsx"

    def test_write_creates_parent_dirs(self, tmp_path: Path) -> None:
        nested = tmp_path / "a" / "b" / "report.xlsx"
        out = write_excel(_sample_dict(), nested)
        assert out.exists()

    def test_write_returns_path(self, tmp_path: Path) -> None:
        out = write_excel(_sample_dict(), tmp_path / "out.xlsx")
        assert isinstance(out, Path)

    def test_written_file_readable(self, tmp_path: Path) -> None:
        out = write_excel(_sample_dict(), tmp_path / "out.xlsx")
        wb = load_workbook(str(out))
        assert "Ranked Summary" in wb.sheetnames
        ws = wb["Ranked Summary"]
        assert ws.max_row == 4  # header + 3 tickers


class TestSignalsFiring:
    def test_signals_firing_rendered_as_comma_string(self) -> None:
        data = _sample_dict(1)
        wb = build_excel_workbook(data)
        ws = wb["Ranked Summary"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        sf_col = headers.index("Signals Firing") + 1
        val = ws.cell(row=2, column=sf_col).value
        assert isinstance(val, str)
        assert "rsi" in val


class TestAgreementFormat:
    def test_agreement_as_fraction(self) -> None:
        data = _sample_dict(1)
        wb = build_excel_workbook(data)
        ws = wb["Ranked Summary"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        agree_col = headers.index("Agreement") + 1
        val = ws.cell(row=2, column=agree_col).value
        assert val == "8/8"
