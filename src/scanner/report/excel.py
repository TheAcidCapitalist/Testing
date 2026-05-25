"""Excel report writer — ranked summary workbook from the canonical dashboard dict.

Consumes the dict produced by ``dashboard_json.build_dashboard_dict``.
Does **not** query DuckDB or re-derive anything.

See ``spec/dashboard-json.md`` for the input contract.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Column spec: (header, dict-path lambda, width hint)
# The lambda receives one ticker dict from data["tickers"].
_COLUMNS: list[tuple[str, callable, int]] = [
    ("Ticker",              lambda t: t["ticker"],                              10),
    ("Exchange",            lambda t: t["exchange"],                            10),
    ("Name",                lambda t: t.get("meta", {}).get("name"),            25),
    ("Direction",           lambda t: t["direction"],                           10),
    ("Combo Score",         lambda t: t["combo_score"],                         12),
    ("Rank Score",          lambda t: t["rank_score"],                          12),
    ("Agreement",           lambda t: _agreement_str(t),                        12),
    ("Signals Firing",      lambda t: ", ".join(t.get("signals_firing", [])),   35),
    ("Vol Confirmation",    lambda t: t.get("vol_confirmation"),                16),
    ("Volume Confirmation", lambda t: t.get("volume_confirmation"),             18),
    ("Days Since Breakout", lambda t: t.get("days_since_breakout"),             18),
    ("Alignment Fraction",  lambda t: t.get("mtf_alignment", {}).get("alignment_fraction"), 16),
    ("Sector",              lambda t: t.get("meta", {}).get("sector"),          18),
    ("Region",              lambda t: t.get("meta", {}).get("region"),          18),
    ("Market Cap (USD)",    lambda t: t.get("meta", {}).get("market_cap_usd"), 18),
]


def build_excel_workbook(data: dict) -> Workbook:
    """Build an openpyxl Workbook from the canonical dashboard dict.

    Parameters
    ----------
    data:
        The dict returned by ``build_dashboard_dict()``.

    Returns
    -------
    openpyxl.Workbook
        A workbook with one sheet ("Ranked Summary"), header row, and one
        data row per scored ticker in rank_score-descending order.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ranked Summary"

    # Header row
    for col_idx, (header, _, width) in enumerate(_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows — order is already rank_score descending from the dict
    for row_idx, ticker in enumerate(data.get("tickers", []), start=2):
        for col_idx, (_, extractor, _) in enumerate(_COLUMNS, start=1):
            value = extractor(ticker)
            # Null-tolerant: write None as blank (openpyxl renders as empty cell)
            ws.cell(row=row_idx, column=col_idx, value=_clean(value))

    return wb


def write_excel(data: dict, path: str | Path) -> Path:
    """Write the canonical dashboard dict to an .xlsx file.

    Creates parent directories if needed.  Returns the resolved ``Path``.
    """
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = build_excel_workbook(data)
    wb.save(str(p))
    return p


# ── Private helpers ──────────────────────────────────────────────────────────


def _agreement_str(t: dict) -> str:
    """Format agreement as 'count/total' e.g. '6/8'."""
    count = t.get("agreement_count", 0)
    total = t.get("n_trade_indicators", 0)
    return f"{count}/{total}"


def _clean(value: object) -> object:
    """Sanitise a value for Excel output.

    - None stays None (renders as blank cell).
    - float('nan') → None (blank, not "nan").
    - str "None" / "nan" → None.
    - Everything else passes through.
    """
    if value is None:
        return None
    if isinstance(value, float):
        import math
        if math.isnan(value):
            return None
    if isinstance(value, str) and value.lower() in ("none", "nan"):
        return None
    return value
